"""Turn Google Forms responses into the client's filled .docx, one per entity.

Google cannot produce the deliverable itself: the templates are the client's
own .docx, with fixed tables and merged cells, and a round trip through Google
Docs does not preserve them. So the forms only collect; the document is still
produced by ``docx_filler``, the same code path the chatbot uses and the one
the template tests cover. Both routes therefore emit byte-identical documents
for identical answers.

    # File > Download > Microsoft Excel from the responses spreadsheet
    python -m app.scripts.from_forms "Etat des lieux - Reponses.xlsx" -o ./sorties

Each sheet of the workbook is one form. Rows are responses; the most recent one
per entity wins unless --all is given.
"""
from __future__ import annotations

import argparse
import csv
import datetime as dt
import os
import re
import sys
import unicodedata
from typing import Any, Dict, List, Optional, Sequence, Tuple

from app.ai.staged import parse_pipe_rows
from app.core.config import settings
from app.pca.blueprint import TEMPLATE_FILES, Question, get_plan
from app.pca.docx_filler import fill_document
from app.scripts.seed import STRUCTURES

# Columns Google adds itself; never a question.
_META = {"horodateur", "timestamp", "adresse e-mail", "email address", "nom d'utilisateur"}

# The dropdown that opens the entity form. Must match STRUCTURE_QUESTION in
# google-forms/Code.gs - it is how a response says which entity it describes.
STRUCTURE_COLUMN = "Structure documentée"

# "Audit Comptable et Financier (ACF)" -> the code in the parentheses.
_CODE_IN_LABEL = re.compile(r"\(([^)]+)\)\s*$")

_UNSAFE = re.compile(r"[^A-Za-z0-9]+")


def _fold(text: str) -> str:
    """Lowercase, unaccented, punctuation-free - for matching only."""
    folded = unicodedata.normalize("NFKD", (text or "").strip().lower())
    folded = "".join(c for c in folded if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]+", " ", folded).strip()


def _slug(text: str) -> str:
    folded = unicodedata.normalize("NFKD", text or "")
    folded = "".join(c for c in folded if not unicodedata.combining(c))
    return _UNSAFE.sub("_", folded).strip("_")[:70] or "entite"


# --------------------------------------------------------------------------- #
# Reading the workbook
# --------------------------------------------------------------------------- #
def _read_xlsx(path: str) -> List[Tuple[str, List[str], List[List[Any]]]]:
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    out = []
    for sheet in book.worksheets:
        rows = [list(r) for r in sheet.iter_rows(values_only=True)]
        if not rows:
            continue
        header = ["" if c is None else str(c).strip() for c in rows[0]]
        body = [r for r in rows[1:] if any(c not in (None, "") for c in r)]
        out.append((sheet.title, header, body))
    book.close()
    return out


def _read_csv(path: str) -> List[Tuple[str, List[str], List[List[Any]]]]:
    with open(path, newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return []
    name = os.path.splitext(os.path.basename(path))[0]
    body = [r for r in rows[1:] if any(c.strip() for c in r)]
    return [(name, [c.strip() for c in rows[0]], body)]


# --------------------------------------------------------------------------- #
# Matching a sheet to a structure
# --------------------------------------------------------------------------- #
# What Google puts in front of the entity name, and what Sheets appends.
_TAB_NOISE = re.compile(
    r"^\s*(?:etat des lieux|copie de)\s*[-–—:]*\s*|\s*\(?reponses?\)?\s*\d*$"
)


def _common_prefix(a: str, b: str) -> int:
    limit = min(len(a), len(b))
    i = 0
    while i < limit and a[i] == b[i]:
        i += 1
    return i


def _by_code(code: str) -> Optional[Tuple[str, str, str, str]]:
    wanted = (code or "").strip().upper()
    for row in STRUCTURES:
        if row[0].upper() == wanted:
            return row
    return None


def structure_from_answer(label: str) -> Optional[Tuple[str, str, str, str]]:
    """Read the entity out of the form's own dropdown answer.

    The single entity form asks which structure it is about, so the answer is
    authoritative - far better than inferring it from a tab name that Excel may
    have truncated. The label is "Nom (CODE)", and the code is matched first
    because two entities can share a name prefix but never a code.
    """
    text = (label or "").strip()
    if not text:
        return None
    found = _CODE_IN_LABEL.search(text)
    if found:
        match = _by_code(found.group(1))
        if match:
            return match
    return _structure_for(text)


def _structure_for(sheet_name: str) -> Optional[Tuple[str, str, str, str]]:
    """Google names each tab after its form: 'État des lieux — <entité>'.

    Matched on a folded prefix, in both directions, because Excel caps a sheet
    name at 31 characters: a real export arrives as "Etat des lieux - Systemes
    d'Inf", with the entity name cut mid-word. Requiring containment found
    nothing at all.
    """
    raw = (sheet_name or "").strip()
    # The per-entity tabs are named "ACF - Audit Comptable et Fin": the code
    # leads precisely because a 31-character cap would cut the name.
    lead = re.match(r"^([A-Za-z0-9]{2,6})\s*[-–—]\s*", raw)
    if lead:
        match = _by_code(lead.group(1))
        if match:
            return match

    candidate = _TAB_NOISE.sub("", _fold(raw)).strip()
    if not candidate:
        return None

    best: Optional[Tuple[int, Tuple[str, str, str, str]]] = None
    for code, name, parent, kind in STRUCTURES:
        target = _fold(name)
        overlap = _common_prefix(candidate, target)
        # The shorter string must be fully consumed - a truncated tab is a
        # prefix of the real name, never a divergence from it.
        if overlap < min(len(candidate), len(target)):
            continue
        # Longest wins, so "Systemes d'Information" beats an entity whose name
        # is merely a prefix of it.
        if best is None or overlap > best[0]:
            best = (overlap, (code, name, parent, kind))
    return best[1] if best else None


# --------------------------------------------------------------------------- #
# Answers
# --------------------------------------------------------------------------- #
def _latest_first(header: Sequence[str], rows: List[List[Any]]) -> List[List[Any]]:
    """Most recent response first, when a timestamp column is present."""
    stamp = next((i for i, h in enumerate(header) if _fold(h) in _META and "mail" not in _fold(h)), None)
    if stamp is None:
        return list(reversed(rows))
    def key(row: List[Any]):
        value = row[stamp] if stamp < len(row) else None
        if isinstance(value, dt.datetime):
            return value
        return dt.datetime.min
    return sorted(rows, key=key, reverse=True)


def build_answers(
    plan: Sequence[Question], header: Sequence[str], row: Sequence[Any]
) -> Tuple[Dict[str, Dict[str, Any]], List[str], List[str]]:
    """Map one response row onto the question plan.

    Returns (answers, blank question labels, warnings).
    """
    by_label = {_fold(h): i for i, h in enumerate(header)}
    answers: Dict[str, Dict[str, Any]] = {}
    blank: List[str] = []
    warnings: List[str] = []

    for question in plan:
        index = by_label.get(_fold(question.label))
        raw = "" if index is None or index >= len(row) else row[index]
        text = ("" if raw is None else str(raw)).strip()

        if index is None:
            warnings.append(f"colonne absente du fichier : « {question.label} »")

        if not text:
            blank.append(question.label)
            continue

        if question.kind == "grid":
            rows = parse_pipe_rows(text, question)
            if not rows:
                # Nothing is dropped: the text goes into the first column and is
                # flagged, because a silent guess on an audit document is worse
                # than a line in a report.
                first = question.columns[0].id if question.columns else "valeur"
                rows = [
                    {c.id: (line.strip() if c.id == first else "") for c in question.columns}
                    for line in text.splitlines() if line.strip()
                ]
                warnings.append(
                    f"« {question.label} » : séparateur « | » absent, "
                    f"{len(rows)} ligne(s) placée(s) dans la première colonne - à relire"
                )
            answers[question.id] = {"rows": rows}
        else:
            answers[question.id] = {"value": text}

    return answers, blank, warnings


# --------------------------------------------------------------------------- #
# Driving
# --------------------------------------------------------------------------- #
def collect(sheets) -> Tuple[Dict[str, List[Tuple[Any, Sequence[str], Sequence[Any]]]], List[str]]:
    """Gather every response in the workbook, bucketed by entity code.

    Two layouts have to work. The single entity form names its entity in a
    dropdown column; the older one-form-per-entity layout only had the tab
    name. The dropdown wins where present, because Excel truncates tab names.

    The per-entity tabs are a derived view of the raw ones, so the same
    response appears twice in an export. Identical rows are collapsed here
    rather than producing the document twice.
    """
    buckets: Dict[str, List[Tuple[Any, Sequence[str], Sequence[Any]]]] = {}
    seen: set = set()
    unmatched: List[str] = []

    for sheet_name, header, rows in sheets:
        if not rows or _fold(sheet_name) == "liens":
            continue

        column = next(
            (i for i, h in enumerate(header) if _fold(h) == _fold(STRUCTURE_COLUMN)), None
        )
        by_tab = _structure_for(sheet_name)

        matched_any = False
        for row in rows:
            structure = None
            if column is not None and column < len(row):
                structure = structure_from_answer(str(row[column] or ""))
            if structure is None:
                structure = by_tab
            if structure is None:
                continue

            matched_any = True
            fingerprint = (structure[0], tuple(str(c) for c in row))
            if fingerprint in seen:
                continue
            seen.add(fingerprint)

            stamp = row[0] if row and isinstance(row[0], dt.datetime) else dt.datetime.min
            buckets.setdefault(structure[0], []).append((stamp, header, row))

        if not matched_any:
            unmatched.append(sheet_name)

    for code in buckets:
        buckets[code].sort(key=lambda item: item[0], reverse=True)
    return buckets, unmatched


def generate(
    path: str, out_dir: str, redacteur: str, keep_all: bool = False
) -> int:
    sheets = _read_csv(path) if path.lower().endswith(".csv") else _read_xlsx(path)
    if not sheets:
        print(f"{path}: aucune feuille exploitable.", file=sys.stderr)
        return 1

    buckets, unmatched = collect(sheets)
    for name in unmatched:
        print(f"  — « {name} » : aucune structure reconnue, ignorée")

    os.makedirs(out_dir, exist_ok=True)
    produced = 0

    for code in sorted(buckets):
        structure = _by_code(code)
        if structure is None:                     # collect() only buckets real codes
            continue
        _code, name, _parent, kind = structure
        plan = get_plan(kind)
        responses = buckets[code] if keep_all else buckets[code][:1]

        for position, (_stamp, header, row) in enumerate(responses):
            answers, blank, warnings = build_answers(plan, header, row)
            if not answers:
                print(f"  — {name} : réponse vide, ignorée")
                continue

            document = fill_document(
                os.path.join(settings.TEMPLATE_DIR, TEMPLATE_FILES[kind]),
                kind,
                answers,
                structure_name=name,
                structure_code=code,
                redacteur=redacteur,
                interview_date=dt.date.today(),
            )

            suffix = f"_{position + 1}" if keep_all and len(responses) > 1 else ""
            target = os.path.join(out_dir, f"Etat_des_lieux_{_slug(name)}{suffix}.docx")
            with open(target, "wb") as handle:
                handle.write(document)
            produced += 1

            filled = len(plan) - len(blank)
            print(f"  ✓ {name} ({code}) — {filled}/{len(plan)} points → {os.path.basename(target)}")
            for warning in warnings:
                print(f"      ! {warning}")
            if blank:
                shown = ", ".join(blank[:4]) + (" …" if len(blank) > 4 else "")
                print(f"      · {len(blank)} point(s) sans réponse : {shown}")

    total = len(buckets)
    print(f"\n{produced} document(s) produit(s) dans {out_dir}"
          + (f" pour {total} entité(s)" if total else ""))
    return 0 if produced else 1


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("responses", help="classeur .xlsx exporté des réponses, ou un .csv")
    parser.add_argument("-o", "--out", default="./sorties", help="dossier de sortie")
    parser.add_argument(
        "--redacteur",
        default=f"Collecte par formulaire - via plateforme {settings.CONSULTING_ORG}",
        help="mention portée dans la fiche de suivi du document",
    )
    parser.add_argument(
        "--all", action="store_true", dest="keep_all",
        help="produire un document par réponse, et non la plus récente seulement",
    )
    args = parser.parse_args(argv)

    if not os.path.exists(args.responses):
        print(f"Fichier introuvable : {args.responses}", file=sys.stderr)
        return 1
    return generate(args.responses, args.out, args.redacteur, args.keep_all)


if __name__ == "__main__":
    sys.exit(main())
